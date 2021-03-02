#!/usr/bin/env python
# coding: utf-8

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import numpy as np
import os
from IPython.display import display
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from IPython.display import display_html
import sys
from enum import Enum
import shutil

class Essay(Enum):
    #[15, 19, 25, 38, 42, 50, 64, 72, 78, 82]
    special_exp = (15, "특별한 경험")
    essay_1 = (19, "물질적 보상")
    essay_2 = (25, "동료와의 관계")
    essay_3 = (38, "업무 환경")
    essay_4 = (42, "사회적 보상")
    essay_5 = (50, "리더와의 관계")
    essay_6 = (64, "성장 환경")
    essay_7 = (72, "정신적 보상")
    essay_8 = (78, "세상과의 관계")
    essay_9 = (82, "통합 환경")
    
    def __init__(self, index, description):
        self.index = index
        self.description = description
        
    def essayIndex():
        return list(map(lambda e: e.index, Essay))
    
    def essayList():
        return list(map(lambda e: e.description, Essay))
        
class ShortAnswer(Enum):
    original_shortAnswer = [16,17,18,20,21,22,23,24,26,27,28,29,30,31,32,33,34,35,36,37,
                             39,40,41,43,44,45,46,47,48,49,51,52,53,54,55,56,57,58,59,60,61,62,63,65,
                             66,67,68,69,70,71,73,74,75,76,77,79,80,81]
    
    original_shortAnswer_Keywords = ['안정감','보상의 공정성','실효적인 복지제도',
                                '동료에 대한 신뢰','뛰어난 동료','사교활동','친밀감', '절친한 친구',
                                '만족감','유연한 출근제도','퇴근의 자율성','업무시간 외 업무지시 금지','휴가사용의 자율성','정규직 비율','정년보장','웃음과 유머','부서간 협업','품의 및 결재','보고방식','회의방식',
                                '존재감','사회적 보상의 중요성 인식','존중감',
                                '리더에 대한 신뢰','언행일치','권한위임','건설적 피드백','노하우 공유','전략적 사고','솔직함',
                                '성취감','성장 중심 육성체계','성장기회','자기계발 기회','도전적 목표수립','육성 중심 평가체계','성과 기반 승진','차별금지','절대평가','다면평가','인사위원회','소통의 투명성','공감 기반 목표수립',
                                '자존감','업무에 대한 흥미','업무를 통한 의미','브랜드파워 및 인지도','꿈의 직장','회사에 대한 자부심','이미지와 사업활동의 일치',
                                '경영진에 대한 신뢰','사회 공동의 이익','선한 영향력','정기적 이익환원','자원봉사 프로그램',
                                '완성감','기업윤리','명확한 방향성']
    
    #응담자 정보 인덱스= [0~14]  id_last_index= 응답자 정보 마지막 인덱스+1
    id_last_index = 4
    
    
    
class CultureIceberg(Enum):
    
    #객관식 총 58문항
    #-----긍정문화   (df 변환 후 인덱스)
    positive = ([0,1,2], "물질적 보상")
    coworker = ([3,4,5,6,7], "동료와의 관계")
    work = ([8,9,10,11,12,13,14,15,16,17,18,19], "업무 환경")
    #-----성과문화
    social = ([20,21,22], "사회적 보상")
    leader = ([23,24,25,26,27,28,29], "리더와의 관계")
    growth = ([30,31,32,33,34,35,36,37,38,39,40,41,42], "성장 환경")
        
    #-----가치문화
    mental = ([43,44,45,46,47,48,49], "정신적 보상")
    world = ([50,51,52,53,54], "세상과의 관계")
    environment = ([55,56,57], "통합 환경")
    
    def __init__(self, index_list, description):
        self.index_list = index_list
        self.description = description
    
    def list():
        return list(map(lambda s: s, CultureIceberg))
    
    def list_description():
        return list(map(lambda s: s.description, CultureIceberg))
    

class GetResponse:
    
    nowRowNumber, lastRowNumber = None, None
    server, df = None, None  #df = dataframe
    company_list_3000 = None
    company_list_dic = {}
    companies = None
    company_list_path, target_path = None, None
    grouped = None

    def __init__(self, json_file_name, spreadsheet_url, db_Sheet, server_Sheet, company_list_path, target_path):
        
        self.json_file_name = json_file_name
        self.spreadsheet_url = spreadsheet_url
        self.db_Sheet = db_Sheet
        self.server_Sheet = server_Sheet
        self.company_list_path = company_list_path
        self.target_path = target_path
        
        self.accessToGspread()
        self.companies = self.df['귀하가 소속해있는 회사명을 기입해주세요'].unique()
        self.grouped = self.df.groupby("귀하가 소속해있는 회사명을 기입해주세요")
        
        self.getCompanyandFolderName(company_list_path)
    
    def accessToGspread(self):
        #------gspread: `잡다문화지수 설문(응답)` 연결하기
        scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive',]
        credentials = ServiceAccountCredentials.from_json_keyfile_name(self.json_file_name, scope)
        gc = gspread.authorize(credentials)
        
        doc = gc.open_by_url(self.spreadsheet_url)
        response_DB = doc.worksheet(self.db_Sheet)
        self.server = doc.worksheet(self.server_Sheet) 
        self.lastRowNumber = int(self.server.acell('A2').value)
        
        #------`잡다문화지수 설문(응답)`의 모든 데이터 불러오기
        gc2 = response_DB.get_all_values()
        self.df = pd.DataFrame(gc2)

        #------`행(=타임스탬프),열(=질문)지정하기
        self.df = self.df.rename(columns=self.df.iloc[0])
        self.df = self.df.drop(self.df.index[0])
        self.df.index=self.df["타임스탬프"]
        self.df = self.df.drop(self.df.columns[0], axis=1)
        
        # [마이다스아이티, 테스트, 한국회사] 더미 데이터 제거
        self.df = self.df.drop(['2021. 1. 26 오후 12:21:08','2021. 1. 28 오후 9:26:13','2021. 2. 17 오전 9:55:13'])
        
        # 총 데이터(현재 gspread)의 마지막 row
        self.nowRowNumber = len(self.df.index)
        
    def showStatus(self):
        #------`잡다문화지수 설문(응답)`에 있는 데이터 요약: 현재 응답 현황
        
        #------response_Status :회사별 현황 테이블
        number_of_Response = len(self.df.index)
        companies = self.df['귀하가 소속해있는 회사명을 기입해주세요'].unique()
        number_of_Companies = len(companies)

        grouped = self.df.groupby('귀하가 소속해있는 회사명을 기입해주세요')
        response_Status = grouped.size()
        response_Status = pd.DataFrame(response_Status, columns=["응답 개수"])
        response_Status.index.name = ""
        response_Status["응답 개수 10개 이상"] = ["●" if num >=10 else "○" for num in response_Status['응답 개수']]
        response_Status = response_Status.sort_values(by='응답 개수' ,ascending=False)
        response_Status["회사별 응답 달성률 (%)"] = (response_Status['응답 개수']/10).map(lambda num: '{0:.1f}%'.format(round(num * 100, 1)))

        #------summary : 전체 응답 현황 테이블
        completed_Company = len(response_Status[response_Status["응답 개수 10개 이상"]=="●"])
        notYet_Company = number_of_Companies - completed_Company
        summary = pd.DataFrame({"총 응답 개수":[number_of_Response], "응답 회사 수":[number_of_Companies],
                                "응답 개수 충족 회사 수":[completed_Company],"응답 개수 미달 회사 수":[notYet_Company]},index=["통계"])
        summary["응답 달성률 (총 기대 응답 수 : 500 개)"] = (summary['총 응답 개수'] / 500).map(lambda num: '{0:.1f}%'.format(round(num * 100, 1)))

        #------기업별 직무 현황 테이블
        html_str=''
        for name, group in grouped:
            tmp = pd.DataFrame(group.iloc[:,3])
            colname = name + " {0}명".format(len(group))
            tmp.columns = [colname]
            tmp = tmp.sort_values(tmp.columns[0])
            tmp_g = tmp.groupby(colname).size()
            tmp = pd.DataFrame(tmp_g, columns=[""])
            html_str+=tmp.to_html()
            
        display(summary)
        display(response_Status)
        display_html(html_str.replace('table','table style="display:inline"'),raw=True)
        
    def storeGspreadCompanyName(self, path): 
        # 이거 사용하지 말고 기업 이름 편집은 엑셀에서 하기
        # path = "C:/Users/ktl0602/Documents/Python Scripts/기업이름_백업.xlsx"

        #----`기업별 폴더명` 시트 생성, path의 excel파일에 3000사 리스트 이미 존재  
        self.company_list_3000 = pd.read_excel(path, sheet_name="기업별 폴더명", usecols="A")
        self.company_list_3000.index += 1
        directory_name_list = []
        
        #----기업번호 생성 : ex) 0001_삼성전자(주) 
        for i, data in self.company_list_3000.iterrows():
            tmp = format(i,'04')
            directory_name = tmp+'_'+data[0]
            directory_name_list.append(directory_name)
        writer = pd.ExcelWriter(path)
        #writer.book = load_workbook(path)
        #writer.sheets = {ws.title: ws for ws in book.worksheets}
        self.company_list_3000['folder_name'] = directory_name_list
        self.company_list_3000.to_excel(writer, sheet_name='기업별 폴더명', index=False)
        writer.save()
        
    def getCompanyandFolderName(self, path): 
        #----기업리스트 DB -> dic{기업명: 폴더명} 생성
        file = pd.read_excel(path, sheet_name="기업별 폴더명")
        company_name = list(file.company_name)
        folder_name = list(file.folder_name)

        for c, f in zip(company_name, folder_name): 
            self.company_list_dic[c] = f
            
    def createDirectory(self, path, generate_path): 
        # 처음에만 사용, 그 이후에 사용 X
        #------기업 폴더 생성 : 기업번호_기업이름 (필요한 정보는 기업리스트.xlsx에 존재)
        cp_list = pd.read_excel(path, sheet_name="기업별 폴더명", usecols="B")
        for i, data in cp_list[:1001].iterrows():
            tmp = generate_path +'/'+ data[0]
            os.makedirs(tmp)
        exception_path = generate_path+"/예외/"
        os.makedirs(exception_path)
        print("폴더 생성 경로 : ", generate_path)
    
    def excelWrite(self, excelpath, company_data, essay):
        # 엑셀 파일 새로 생성 (덮어쓰기)
        writer = pd.ExcelWriter(excelpath, mode='W')
        company_data.to_excel(writer, sheet_name='RawData')
        essay.to_excel(writer, sheet_name='Essay')
        writer.save()
        
    def excelWriteToExistFile(self, excelpath, company_data, essay):
        # 엑셀 파일에 이어 쓰기
        writer = pd.ExcelWriter(excelpath, engine='openpyxl')
        writer.book = load_workbook(excelpath)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        company_data.to_excel(writer,sheet_name='RawData', startrow=writer.sheets['RawData'].max_row, header= False)
        essay.to_excel(writer,sheet_name='Essay', startrow=writer.sheets['Essay'].max_row, header= False)
        writer.save()  
        
    def editExcelWrite(self, edit_excelpath, company_data):
        writer = pd.ExcelWriter(edit_excelpath, engine='openpyxl')
        writer.book = load_workbook(edit_excelpath)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        company_data.to_excel(writer,sheet_name='RawData')
        writer.save()  
        
    def processAllData(self, path): 
        # 기업별 excel파일 모두 생성하므로 지역변수 path 사용, target_path(클래스 변수) 사용 X
        #------모든 응답 데이터를 처리해서 각 기업별 엑셀 파일 생성
        
        for company_name, company_data in self.grouped:
            essay = company_data.iloc[:, Essay.essayIndex()]
            company_folder_name = self.company_list_dic.get(company_name)
            if company_folder_name != None:
                #print(company_folder_name)
                directory_path = path + "/{0}".format(company_folder_name)
                excelpath = directory_path + "/{0}.xlsx".format(company_folder_name)
                
                if not os.path.exists(directory_path):
                    os.makedirs(directory_path)
                    print("folder is generated ", directory_path)
                
                edit_excelpath = directory_path + "/{0}_문화분석.xlsx".format(company_name)
                shutil.copy("C:/Users/ktl0602/Documents/Python Scripts/문화분석.xlsx", edit_excelpath)
                
                print("file is generated ", company_folder_name)
                self.excelWrite(excelpath, company_data, essay)
                self.editExcelWrite(edit_excelpath, company_data)
                
            else:
                excp_path = path + "/예외"
                if not os.path.exists(excp_path):
                    os.makedirs(excp_path)
                excelpath = excp_path +"/{0}.xlsx".format(company_name)
                self.excelWrite(excelpath, company_data, essay)
                print("file is generated",excelpath)
                
        self.server.update('A2', self.nowRowNumber)
        
        print("처리한 데이터 개수: ", len(self.df), " 생성된 엑셀 파일 개수(회사 수): ", len(self.companies))
        print("생성된 엑셀 파일 경로: ",path )
           
    def processAddedData(self):
        #-----추가된 데이터 처리, 예외 폴더가 있어야 동작
        added_df = self.df[self.lastRowNumber:]
        grouped = self.df[self.lastRowNumber:].groupby("귀하가 소속해있는 회사명을 기입해주세요")
        for company_name, company_data in grouped:
            essay = company_data.iloc[:, Essay.essayIndex() ]
            company_folder_name = self.company_list_dic.get(company_name)
            if company_folder_name != None:
                directory_path = self.target_path + "/{0}".format(company_folder_name)
                excelpath = directory_path + "/{0}.xlsx".format(company_folder_name)

                if not os.path.exists(excelpath): 
                    # 폴더는 1000번까지만 생성했기 때문에 1000번 이후의 기업 데이터가 들어오면 폴더부터 생성
                    if not os.path.exists(directory_path):
                        os.makedirs(directory_path)
                        print("folder is generated ", directory_path)
                        edit_excelpath = directory_path + "/{0}_문화분석.xlsx".format(company_name)
                        shutil.copy("C:/Users/ktl0602/Documents/Python Scripts/문화분석.xlsx", edit_excelpath)
                        self.editExcelWrite(edit_excelpath, company_data)
                        
                    print("file is generated ", company_folder_name)
                    self.excelWrite(excelpath, company_data, essay)   
                    
                else: 
                    print("excists",directory_path)
                    self.excelWriteToExistFile(excelpath, company_data, essay)

            else: #3000사 리스트에 없는 응답들
                excelpath = self.target_path +'/예외/{0}.xlsx'.format(company_name) 
                if not os.path.exists(excelpath):
                    print("file is generated",excelpath)
                    self.excelWrite(excelpath, company_data, essay)                   
                else:
                    print("excists",excelpath)
                    self.excelWriteToExistFile(excelpath, company_data, essay)

        self.server.update('A2', self.nowRowNumber)

        print("처리한 데이터 개수: ", len(added_df) , " 수정된(생성된) 엑셀 파일 개수(회사 수): ", len(grouped))
        print("생성된 엑셀 파일 경로: ", self.target_path )
        
    def positiveResponse(self, path):
        # 문화분석리포트용 rawData 생성 함수
        for company_name, company_data in self.grouped:
            if len(company_data) < 10: continue
            #if company_name !="(주)카카오": continue
            company_data = company_data.sort_index(ascending=True)
            print(company_name,"결과")
            
            #excelpath = r"C:\Users\ktl0602\Documents\Python Scripts\문화분석리포트\{0}.xlsx".format(company_name)
            excelpath = path + '/{0}.xlsx'.format(company_name)
            shutil.copy("C:/Users/ktl0602/Documents/Python Scripts/기업진단리포트.xlsx", excelpath)
            
            #excelpath = path +'/test.xlsx'
            writer = pd.ExcelWriter(excelpath, engine='openpyxl')
            writer.book = load_workbook(excelpath)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            

            #---[응답자 정보] sheet
            identifier = company_data.iloc[:,: ShortAnswer.id_last_index.value]
            display(identifier)
            
            identifier.to_excel(writer,sheet_name='응답자 정보')

            company_data_short_answer = company_data.iloc[:,ShortAnswer.original_shortAnswer.value]
            show_company_data = company_data_short_answer.apply(lambda x: x.value_counts()).T
            show_company_data = pd.DataFrame(show_company_data, columns=['그렇다','그렇지 않다','매우 그렇다','매우 그렇지 않다','모르겠다','보통이다','아니다']).fillna(0)
            short_answer_num = len(company_data_short_answer)
            show_company_data['키워드'] = ShortAnswer.original_shortAnswer_Keywords.value
            #show_company_data['비율(%)'] = ((show_company_data['그렇다'] + show_company_data['매우 그렇다'] + show_company_data['보통이다'])  / len(company_data) *100).map(lambda num: '{0:.1f}%'.format(num,0))
            show_company_data['비율(%)'] = ((show_company_data['그렇다'] + show_company_data['매우 그렇다'] + show_company_data['보통이다'])  / len(company_data) *100)
            display(show_company_data)
            
            Total = show_company_data['그렇다'].sum() + show_company_data['매우 그렇다'].sum() +show_company_data['보통이다'].sum()
            
            total_index = "전체 긍정 응답 비율"
            total = (Total/(len(ShortAnswer.original_shortAnswer.value)*len(company_data)))*100
            total_res = '{0:.1f}%'.format((Total/(len(ShortAnswer.original_shortAnswer.value)*len(company_data)))*100)

            print(total_index,' : ' ,total_res)
            culture_list={}
            culture_list[total_index] = total
            

             #---[문항별 수치] sheet
            show_company_data.to_excel(writer, sheet_name='문항별 수치')


            for i, culture in enumerate(CultureIceberg.list()):
                division = show_company_data.iloc[culture.index_list,:]
                Total = division['그렇다'].sum() + division['매우 그렇다'].sum() +division['보통이다'].sum();

                table_index = '{0}'.format(culture.description)
                #ratio = '{0:.1f}'.format((Total/(len(division)*short_answer_num))*100)
                ratio = (Total/(len(division)*short_answer_num))*100
                tmp =table_index +' : '+ str(ratio)
                print(table_index,' : ', ratio)
                display(division)
                culture_list[table_index] = ratio

                ws =  writer.book['문항별 수치']
                ws.append([''])
                ws.append([tmp])
                division.to_excel(writer,sheet_name='문항별 수치', startrow=writer.sheets['문항별 수치'].max_row, header= True)
                ws.append([''])

            #---[문화별 Essay] sheet
            essay = company_data.iloc[:,Essay.essayIndex()]
            essay = essay.T.reset_index().rename(columns={"index": "문항"})
            essay.index = Essay.essayList()
            essay = essay.T
            display(essay)
            essay.to_excel(writer,sheet_name='문화별 Essay')
            
            #---[데이터] sheet
            
            years  = identifier.groupby('귀하가 본 회사에서 근무한 기간을 입력해주세요 [근무기간]')
            positions = identifier.groupby('귀하가 담당하고 있는 직군을 선택해주세요.')
            years_Status = years.size()
            positions_Status = positions.size()
            years_Status = pd.DataFrame(years_Status, columns=["근무기간"])
            #display(years_Status)
            #display(positions_Status)
            
            
            basic_data = pd.DataFrame({"기업명":[company_name],"참여인원": [len(company_data)],
                                "시작일":[company_data.index[0]],"종료일":[company_data.index[-1]]},index=["기본 정보"])
            basic_data = basic_data.T
            basic_data.to_excel(writer,sheet_name='데이터')

            display(basic_data)
            
            p = (float(culture_list["물질적 보상"])+float(culture_list["동료와의 관계"])+float(culture_list["업무 환경"]))/3.0
            a = (float(culture_list["사회적 보상"])+float(culture_list["리더와의 관계"])+float(culture_list["성장 환경"]))/3.0
            v = (float(culture_list["정신적 보상"])+float(culture_list["세상과의 관계"])+float(culture_list["통합 환경"]))/3.0
            
            sex_data = pd.DataFrame({"남성":["=COUNTIF('응답자 정보'!B:B,A7)"],"여성":["=COUNTIF('응답자 정보'!B:B,A8)"]},
                                    index=["성별"])
            years_data = pd.DataFrame({"3~5년":["=COUNTIF('응답자 정보'!D:D,A10)"], "6~8년":["=COUNTIF('응답자 정보'!D:D,A11)"],
                                       "9~11년":["=COUNTIF('응답자 정보'!D:D,A12)"],"12년 이상":["=COUNTIF('응답자 정보'!D:D,A13)"]},
                                      index=["근속연수"])
            position_data = pd.DataFrame({"영업":["=COUNTIF('응답자 정보'!E:E,A15)"],"연구/개발":["=COUNTIF('응답자 정보'!E:E,A16)"],
                                "IT서비스":["=COUNTIF('응답자 정보'!E:E,A17)"],"경영기획/지원":["=COUNTIF('응답자 정보'!E:E,A18)"],
                                "생산/유통/품질":["=COUNTIF('응답자 정보'!E:E,A19)"],
                                "홍보/마케팅":["=COUNTIF('응답자 정보'!E:E,A20)"],
                                "서비스/고객지원":["=COUNTIF('응답자 정보'!E:E,A21)"],
                                "의료/보건":["=COUNTIF('응답자 정보'!E:E,A22)"],
                                "건설엔지니어":["=COUNTIF('응답자 정보'!E:E,A23)"],
                                "금융/보험":["=COUNTIF('응답자 정보'!E:E,A24)"],
                                "디자인": ["=COUNTIF('응답자 정보'!E:E,A25)"]},index=["직군"])
            
            
            sex_data.T.to_excel(writer,sheet_name='데이터',startrow=writer.sheets['데이터'].max_row, header= True)
            years_data.T.to_excel(writer,sheet_name='데이터',startrow=writer.sheets['데이터'].max_row, header= True)
            position_data.T.to_excel(writer,sheet_name='데이터',startrow=writer.sheets['데이터'].max_row, header= True)
            
            culture_list =  pd.DataFrame( culture_list ,index=['긍정응답 비율(%)']).T
            culture_list = culture_list.round(0).astype(int)
            
            display(culture_list)
            
            
            culture_list_sorted = culture_list.sort_values(by=['긍정응답 비율(%)'], ascending=False)
            top_3 = culture_list_sorted.head(3)
            top_3.columns = ['상위 긍정 응답 문화']
            bottom_3 = culture_list_sorted.tail(3)
            bottom_3.columns = ['하위 긍정 응답 문화']
            
            display(top_3)
            display(bottom_3)
           
            culture_iceberg = pd.DataFrame({"긍정문화": [p] ,"성과문화":[a],"가치문화":[v]},index=["문화 지수 요약"])
            culture_iceberg = culture_iceberg.round(0).astype(int)
            display(culture_iceberg.T)
            culture_iceberg =culture_iceberg.T
            culture_iceberg.to_excel(writer,sheet_name='데이터',startrow=writer.sheets['데이터'].max_row, header= True)
            top_3.to_excel(writer,sheet_name='데이터',startrow=writer.sheets['데이터'].max_row, header= True)
            bottom_3.to_excel(writer,sheet_name='데이터',startrow=writer.sheets['데이터'].max_row, header= True)
            culture_list.to_excel(writer,sheet_name='데이터',startrow=writer.sheets['데이터'].max_row, header= True)
            writer.save() 


if __name__ == '__main__':   
    
    json_file_name = 'C:/Users/ktl0602/Downloads/maximal-storm-303606-00a7c52333a6.json'
    spreadsheet_url = 'https://docs.google.com/spreadsheets/d/14SvAKSzEbbP0pSASUmuCdRtqP7pKcmf_oT9nklGKGg0/edit#gid=483125835'
    db_Sheet = '설문지 응답 시트1'
    server_Sheet = 'Server'
    company_list_path = 'C:/Users/ktl0602/Documents/Python Scripts/기업이름DB.xlsx'
    #target_path= '//midasfile4/900_Shared_Folder/Team/전략기획팀/협력직/기업데이터 수집/문화지수응답'
    target_path= '//midasfile4/900_Shared_Folder/Team/전략기획팀/협력직/문화지수 설문 수집/설문 데이터'
    #target_path = 'C:/Users/ktl0602/Documents/Results/테스트'
    ## path 입력은 \ (X), / (O)
    
    conn = GetResponse(json_file_name, spreadsheet_url, db_Sheet, server_Sheet, company_list_path, target_path)
    conn.showStatus()
    #conn.positiveResponse()
    #conn.processAddedData()
    #conn.processAllData(target_path)

